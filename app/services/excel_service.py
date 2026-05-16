from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.ai.sql.dataframe_analyzer import detect_column_type


class ExcelProcessingError(ValueError):
    pass


def _sheet_preview(sheet, max_preview_rows: int = 50) -> tuple[list[str], list[dict[str, Any]], int, dict[str, str]]:
    iterator = sheet.iter_rows(values_only=True)
    header = next(iterator, None)
    if not header:
        raise ExcelProcessingError("Sheet has no rows")

    columns = [str(col) if col is not None else "" for col in header]
    if not any(columns):
        raise ExcelProcessingError("Sheet has empty header")

    rows: list[dict[str, Any]] = []
    total_rows = 0
    for values in iterator:
        total_rows += 1
        if len(rows) < max_preview_rows:
            row = {}
            for index, column in enumerate(columns):
                cell = values[index] if index < len(values) else None
                row[column] = cell
            rows.append(row)

    column_types = {column: detect_column_type([row.get(column) for row in rows]) for column in columns}
    return columns, rows, total_rows, column_types


def parse_excel_bytes(content: bytes, selected_sheet: str | None = None, max_preview_rows: int = 50) -> dict[str, Any]:
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelProcessingError(f"Corrupted or unreadable Excel file: {exc}")

    sheet_names = workbook.sheetnames
    if not sheet_names:
        workbook.close()
        raise ExcelProcessingError("Excel file contains no sheets")

    target_sheet_name = selected_sheet if selected_sheet in sheet_names else sheet_names[0]
    sheet = workbook[target_sheet_name]

    columns, rows, total_rows, column_types = _sheet_preview(sheet, max_preview_rows=max_preview_rows)
    workbook.close()

    return {
        "sheet_names": sheet_names,
        "selected_sheet": target_sheet_name,
        "columns": columns,
        "preview_rows": rows,
        "total_rows": total_rows,
        "column_types": column_types,
    }
