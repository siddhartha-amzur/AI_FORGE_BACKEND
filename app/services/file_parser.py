import asyncio
import base64
import csv
from io import StringIO
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader

from app.models.attachment import Attachment


MAX_TEXT_CHARS = 8000


def _trim_text(text: str) -> str:
    cleaned = text.strip()
    if len(cleaned) <= MAX_TEXT_CHARS:
        return cleaned
    return cleaned[:MAX_TEXT_CHARS].rstrip() + "\n...[truncated]"


def _read_text_file(file_path: str) -> str:
    return _trim_text(Path(file_path).read_text(encoding="utf-8", errors="replace"))


def _read_csv_file(file_path: str) -> str:
    rows: List[str] = []
    with Path(file_path).open("r", encoding="utf-8", errors="replace", newline="") as csv_file:
        reader = csv.reader(csv_file)
        for index, row in enumerate(reader, start=1):
            rows.append(f"Row {index}: " + " | ".join(row))
            if len("\n".join(rows)) >= MAX_TEXT_CHARS:
                break
    return _trim_text("\n".join(rows))


def _read_xlsx_file(file_path: str) -> str:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    lines: List[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"Sheet: {sheet.title}")
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else str(value) for value in row]
            lines.append(f"Row {row_index}: " + " | ".join(values))
            if len("\n".join(lines)) >= MAX_TEXT_CHARS:
                workbook.close()
                return _trim_text("\n".join(lines))
    workbook.close()
    return _trim_text("\n".join(lines))


def _read_pdf_file(file_path: str) -> str:
    try:
        reader = PdfReader(file_path)
        pages = []
        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"Page {index}:\n{text.strip()}")
            if len("\n\n".join(pages)) >= MAX_TEXT_CHARS:
                break
        if not pages:
            return "PDF uploaded successfully, but no extractable text was found in the document."
        return _trim_text("\n\n".join(pages))
    except Exception as exc:
        return f"PDF could not be read safely: {exc}"


def _image_to_data_url(file_path: str, mime_type: str) -> str:
    encoded = base64.b64encode(Path(file_path).read_bytes()).decode("utf-8")
    return f"data:{mime_type};base64,{encoded}"


async def build_attachment_prompt_parts(attachments: List[Attachment]) -> Tuple[str, List[dict]]:
    sections: List[str] = []
    image_parts: List[dict] = []

    print("[file_parser] attachment count:", len(attachments))

    for attachment in attachments:
        print("[file_parser] parsing attachment:", attachment.id, attachment.original_filename)
        print("[file_parser] attachment path:", attachment.file_path)
        suffix = Path(attachment.original_filename).suffix.lower()
        file_path = Path(attachment.file_path)

        if not file_path.exists():
            print("[file_parser] missing file on disk:", attachment.file_path)
            sections.append(
                f"Attachment: {attachment.original_filename} ({attachment.mime_type})\nStored file is missing on disk, so no content could be parsed."
            )
            continue

        if attachment.mime_type.startswith("image/"):
            sections.append(
                f"Attachment: {attachment.original_filename} ({attachment.mime_type}). Use the attached image to answer image-related questions."
            )
            data_url = await asyncio.to_thread(_image_to_data_url, attachment.file_path, attachment.mime_type)
            image_parts.append({"type": "image_url", "image_url": {"url": data_url}})
            print("[file_parser] image converted to data url for:", attachment.original_filename)
            continue

        if attachment.mime_type.startswith("video/"):
            sections.append(
                f"Attachment: {attachment.original_filename} ({attachment.mime_type}, {attachment.file_size} bytes). Video understanding is not implemented yet, so only file metadata is available."
            )
            print("[file_parser] video metadata only for:", attachment.original_filename)
            continue

        if suffix == ".pdf":
            parsed_text = await asyncio.to_thread(_read_pdf_file, attachment.file_path)
        elif suffix in {".csv"}:
            parsed_text = await asyncio.to_thread(_read_csv_file, attachment.file_path)
        elif suffix in {".xlsx"}:
            parsed_text = await asyncio.to_thread(_read_xlsx_file, attachment.file_path)
        else:
            parsed_text = await asyncio.to_thread(_read_text_file, attachment.file_path)

        preview = parsed_text[:500]
        print("[file_parser] parsed content preview:", preview)
        if not parsed_text.strip():
            parsed_text = "File was uploaded successfully but no readable text content was extracted."
            print("[file_parser] empty parsed content for:", attachment.original_filename)

        sections.append(
            f"Attachment: {attachment.original_filename} ({attachment.mime_type})\n{parsed_text}"
        )

    if not sections:
        print("[file_parser] no attachment sections generated")
        return "", image_parts

    attachment_context = "Attached file content:\n" + "\n\n".join(sections)
    print("[file_parser] attachment context length:", len(attachment_context))
    return attachment_context, image_parts